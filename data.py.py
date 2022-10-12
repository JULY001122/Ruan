import openpyxl
import random
import pandas as pd
classmaxcount = 0
a = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
cname=['a','b','c','d','e']
minp = 6
maxp = 8
minc = 17
maxc = 20
lowgpa = 10
xlsx = openpyxl.Workbook()
sheet = xlsx.active
s1 = '学生'
s2 = '课程'


for i in range (0, 5):
    sheet.cell(row=i * 22 + 1, column=1, value= s2 + cname[i])
    sheet.cell(row=i * 22 + 22, column=1, value='绩点')

for j in range(0,5):
    for i in range(1, 91):
        sheet.cell(row=j * 22 + 1, column=i + 1, value=s1 + str(i))
        sheet.cell(row=j * 22 + 22, column=i + 1, value=0)

for j in range(0,5):
    for i in range(1, 21):
        sheet.cell(row=i + j * 22 + 1, column=1, value=i)

for n in range(0,5):
    for i in range(2, 22):
        for j in range(2, 92):
            sheet.cell(row=i+ n * 22 , column=j, value=1)

for z in range (0,5):
    classcount = [0] * 20
    arr = [0] * 91
    b = [0] * 90
    for i in range(0, 90):
        b[i] = i + 1
    i = random.randint(minp, maxp)
    coli = random.sample(b, i)##随机选缺席16次以上的学生id
    # print(coli)
    for m in range(1, i + 1):
            gpa = round(random.uniform(2.50, 3.20), 3)  ##随机产生较低绩点数
            sheet.cell(row=z * 22 + 22, column=coli[m-1] + 1, value=gpa)
            count = random.randint(17, 20)  ##随机选择他的缺课次数
           ## print(count)
            col = random.sample(a, count)  ##随机选择他缺第x次课
           ## print(col)
            for n in range(1, count + 1):
                sheet.cell(row=col[n-1]+z*22+1, column=coli[m-1] + 1, value=0)

    i =  lowgpa- i  ##较低绩点为固定的人数count
    for m in range(1,91): ##赋予剩下的学生绩点
        cell_obj = sheet.cell(row=z * 22 + 22, column= m+1)
        if cell_obj.value==0 :
            if i < 0 :
                gpa = round(random.uniform(3.20, 3.92), 3)
                sheet.cell(row=z * 22 + 22, column=m + 1, value=gpa)
            else:
                gpa = round(random.uniform(2.00, 3.20), 3)  ##随机产生较低绩点数
                sheet.cell(row=z * 22 + 22, column=m + 1, value=gpa)
                i = i -1


    def classdata(i,min,max):
        for m in range(1, i + 1):
            stuid = random.randint(1, 90)
            flag = [0] * 20
            if arr[stuid] == 0 :  ##没被选
                arr[stuid] = 1
                count = random.randint(min, max)  ##随机选择他的缺课次数
                # print(count ,"count")
                # print(stuid,"stuid")
                for j in range(1,count+1):
                    class_id = random.randint(1, 20)
                    if classcount[class_id-1]==3:
                        while 1:
                            class_id = random.randint(1, 20)
                            if classcount[class_id - 1] < 3 and flag[class_id - 1] == 0:  ##每门课除经常缺席学生外，只有0-3名学生缺席
                                classcount[class_id - 1] += 1
                                flag[class_id - 1] = 1      ##标记没来的课程号
                                sheet.cell(row=class_id + z*22+1, column=stuid + 1, value=0)
                                break
                    else:
                        if flag[class_id - 1] == 0:
                            classcount[class_id - 1] += 1
                            flag[class_id - 1] = 1
                            sheet.cell(row=class_id + z*22+1, column=stuid + 1, value=0)
            else:
                m=m-1
    x = random.randint(2, 3)
    y = random.randint(2, 4)
    classdata(x,8,15)
    classdata(y,3,8)
    for i in range(1,21):
        if classcount[i - 1] <= 3:
            stuid = random.randint(1, 90)
            if arr[stuid] == 0:  ##该学生没在经常逃课的学生名单内
                arr[stuid] = 1
                classcount[i - 1] += 1
                sheet.cell(row=i + z*22+1, column=stuid + 1, value=0)
            else:
                while arr[stuid]==1:
                    stuid = random.randint(1, 90)
                arr[stuid] = 1
                classcount[i - 1] += 1
                sheet.cell(row=i + z*22+1, column=stuid + 1, value=0)
xlsx.save('sample.xlsx')
##把需要转化的xlsx文件放在pycharm项目文件目录里
def xlsx_to_csv_pd():
    data_xls = pd.read_excel('sample.xlsx', index_col=0)  # 输入xlsx文件名
    data_xls.to_csv('sample.csv', encoding='utf-8')  # 输出csv文件名
if __name__ == '__main__':
    xlsx_to_csv_pd()